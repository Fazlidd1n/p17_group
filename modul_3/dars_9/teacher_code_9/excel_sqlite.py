import sqlite3

import openpyxl


def read_excel(filename):
    path = filename
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    data = dict()
    for i in range(1,max_row+1):
        eng = sheet_obj.cell(row=i, column=1).value
        uzb = sheet_obj.cell(row=i, column=2).value
        data[eng] = uzb
    return data


class DB:
    con = sqlite3.connect("translate.sqlite")
    cur = con.cursor()
    translate_table = """
    create table if not exists dictionary(
        id integer primary key autoincrement ,
        eng varchar(255),
        uzb varchar(255),
        unit integer ,
        book integer 
    )
    """
    cur.execute(translate_table)
    con.commit()


    def write(self, data : dict):
        query = """
        insert into dictionary(eng , uzb)
        values (?,?)
        """
        for param in data.items():
            self.cur.execute(query , param)
        self.con.commit()
        return "Success"

DB().write(read_excel("Book2_unit10.xlsx"))







